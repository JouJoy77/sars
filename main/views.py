from django.urls import reverse
import openpyxl
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from django.shortcuts import render

from users.models import User

# вьюшка для главной страницы
def index(request):
    users = User.objects.order_by('-points')
    table_data = [
        {'rank': rank, 'snils': user.snils, 'points': user.points}
        for rank, user in enumerate(users, start=1)
    ]
    context = {
        'Name': 'Главная страница',
        'table_data': table_data,
        'download_url': reverse('download_excel')
    }
    return render(request, 'index.html', context)

# Вьюшка для загрузки excel
def download_excel(request):
    users = User.objects.order_by('-points')
    # Создание нового Excel-документа
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # Заголовки столбцов
    headers = ['Rank', 'SNILS', 'Points']
    for col_num, header in enumerate(headers, start=1):
        column_letter = get_column_letter(col_num)
        sheet[column_letter + '1'] = header
        sheet[column_letter + '1'].font = Font(bold=True)
    # Данные таблицы
    for rank, user in enumerate(users, start=2):
        sheet.cell(row=rank, column=1, value=rank - 1)
        sheet.cell(row=rank, column=2, value=user.snils)
        sheet.cell(row=rank, column=3, value=user.points)
    # Сохранение Excel-документа в памяти
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="rating.xlsx"'
    workbook.save(response)
    return response
